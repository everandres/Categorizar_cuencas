# -*- coding: utf-8 -*-
"""
Created on Fri Jul  1 14:30:31 2022

@author: easalazarm
"""

from openpyxl import load_workbook

precipitacion = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Direccion General\Maximas\Estaciones_Temp_Max.xlsx")
catalogo = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Direccion General\Maximas\Temp_Max_Hist.xlsx")

h_catalogo = catalogo['Hoja1']
h_precipitacion = precipitacion["Hoja1"]

def subzonas(hidro, prec):
    for n in prec['D']:
        for m in hidro['A']:
            if n.value == m.value:
                prec.cell(row = n.row, column = 9).value = hidro.cell(row = m.row, column = 2).value #Area hidrografica
                prec.cell(row = n.row, column = 10).value = hidro.cell(row = m.row, column = 3).value #Zona hidrográfica
                prec.cell(row = n.row, column = 11).value = hidro.cell(row = m.row, column = 4).value #Subzona hidrográfica
                
    
    return precipitacion.save(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Direccion General\Maximas\archivo_temp_organizado.xlsx")

print(subzonas(h_catalogo, h_precipitacion))