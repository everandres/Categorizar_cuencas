# -*- coding: utf-8 -*-
"""
Created on Fri May  6 10:46:14 2022

@author: easalazarm
"""

from openpyxl import load_workbook

catalogo = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Codigos\hidrografia\CATALOGO_IDEAM.xlsx")
precipitacion = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Codigos\hidrografia\precipitacion.xlsx")

h_catalogo = catalogo['CNE']
h_precipitacion = precipitacion["precipitacion"]

def subzonas(hidro, prec):
    for n in prec['D']:
        for m in hidro['B']:
            if n.value == m.value:
                prec.cell(row = n.row, column = 5).value = hidro.cell(row = m.row, column = 14).value #Area hidrografica
                prec.cell(row = n.row, column = 40).value = hidro.cell(row = m.row, column = 15).value #Zona hidrográfica
                prec.cell(row = n.row, column = 46).value = hidro.cell(row = m.row, column = 19).value #Subzona hidrográfica
                
    
    return precipitacion.save(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Codigos\hidrografia\precipitacion_cuencas.xlsx")

print(subzonas(h_catalogo, h_precipitacion))